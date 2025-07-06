from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection, Font, Border, Fill, Alignment
from datetime import datetime
import copy

def copiar_formatacao(cell_origem, cell_destino):
    
    if cell_origem.has_style:
        cell_destino.font = copy.copy(cell_origem.font)
        cell_destino.border = copy.copy(cell_origem.border)
        cell_destino.fill = copy.copy(cell_origem.fill)
        cell_destino.number_format = copy.copy(cell_origem.number_format)
        cell_destino.protection = copy.copy(cell_origem.protection)
        cell_destino.alignment = copy.copy(cell_origem.alignment)


mes = 7  
ano = 2025
arquivo_origem = 'tabela1.xlsx'
arquivo_saida = f'tabela_mes{mes:02d}_{ano}.xlsx'


wb = load_workbook(arquivo_origem)
aba_modelo = wb.active


if not any(aba_modelo.iter_rows()):
    print("ATENÇÃO: A aba modelo parece estar vazia!")
else:
    print(f"Modelo carregado com sucesso (dimensões: {aba_modelo.max_row} linhas × {aba_modelo.max_column} colunas)")


aba_modelo.title = "MODELO_TEMP"  

for dia in range(1, 32):
    try:
        data = datetime(ano, mes, dia)
        nome_aba = data.strftime('%d-%m-%Y')
        
        print(f"\nCriando aba: {nome_aba}...")
        
        
        nova_aba = wb.create_sheet(title=nome_aba)
        
      
        for row in aba_modelo.iter_rows():
            for cell in row:
                nova_cell = nova_aba[cell.coordinate]
                nova_cell.value = cell.value
                copiar_formatacao(cell, nova_cell)
       
        for col in aba_modelo.columns:
            col_letter = get_column_letter(col[0].column)
            nova_aba.column_dimensions[col_letter].width = aba_modelo.column_dimensions[col_letter].width
        
        for row in range(1, aba_modelo.max_row + 1):
            nova_aba.row_dimensions[row].height = aba_modelo.row_dimensions[row].height
        
       
        for merge in aba_modelo.merged_cells.ranges:
            nova_aba.merge_cells(str(merge))
        
        print(f"Aba {nome_aba} criada com sucesso!")
    
    except ValueError:
        print(f" Dia {dia} não existe no mês {mes}, pulando...")
        continue


wb.remove(aba_modelo)
wb.save(arquivo_saida)

print(f"\nArquivo '{arquivo_saida}' gerado com sucesso!")
print(f"Total de abas criadas: {len(wb.sheetnames)}")